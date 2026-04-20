#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "openpyxl",
# ]
# ///
"""Automatically fill a 研究業務日誌 Excel workbook."""

from __future__ import annotations

import argparse
import sys
from datetime import time
from pathlib import Path

OUTPUT_PREFIX = "YANG_FAN_"
BUSHO = "コンテンツ科学研究系"
SHIMEI = "YANG Fan"
SHIGYO_JIKOKU = time(9, 0)
SHUGYO_JIKOKU = time(17, 45)
KYUKEI_JIKAN = time(1, 0)
WEEKEND = {"土", "日"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill a 研究業務日誌 .xlsx file and save the result next to the input file.")
    parser.add_argument("input_file", help="Path to the input .xlsx file")
    return parser.parse_args()


def build_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{OUTPUT_PREFIX}{input_path.name}")


def fill_nisshi(input_path: Path) -> Path:
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Expected an .xlsx file, got: {input_path.name}")

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Missing dependency: openpyxl. Run this script through uv, for example `uv run scripts/fill_research_worklog.py <input.xlsx>`.") from exc

    output_path = build_output_path(input_path)

    wb = load_workbook(input_path)
    ws = wb.active
    wb_data = load_workbook(input_path, data_only=True)
    ws_data = wb_data.active

    ws["T4"] = BUSHO
    ws["T6"] = SHIMEI

    for row_idx in range(9, 40):
        youbi_val = ws_data.cell(row=row_idx, column=2).value
        biko_val = ws_data.cell(row=row_idx, column=22).value

        if youbi_val in WEEKEND:
            continue

        if biko_val and str(biko_val).strip():
            continue

        ws.cell(row=row_idx, column=3).value = SHIGYO_JIKOKU
        ws.cell(row=row_idx, column=4).value = SHUGYO_JIKOKU
        ws.cell(row=row_idx, column=5).value = KYUKEI_JIKAN
        ws.cell(row=row_idx, column=7).value = "■"

    ws["V42"] = "未提出"
    wb.save(output_path)
    return output_path


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_file)
    try:
        output_path = fill_nisshi(input_path)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(output_path.as_posix())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
